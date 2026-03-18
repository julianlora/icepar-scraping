from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def debe_eliminar(valor: object) -> bool:
    if valor is None:
        return False
    if isinstance(valor, str):
        valor_limpio = valor.strip()
        return valor_limpio in {"0", "-"}
    if isinstance(valor, (int, float)):
        return valor == 0
    return False


def procesar_excel_a_csv(archivo_entrada: Path, archivo_salida: Path, hoja: str | None = None) -> tuple[int, int]:
    workbook = load_workbook(filename=archivo_entrada, read_only=True, data_only=True)

    if hoja:
        if hoja not in workbook.sheetnames:
            raise ValueError(f"La hoja '{hoja}' no existe en el archivo")
        worksheet = workbook[hoja]
    else:
        worksheet = workbook.active

    filas_leidas = 0
    filas_escritas = 0

    with archivo_salida.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)
        for indice_fila, fila in enumerate(worksheet.iter_rows(values_only=True), start=1):
            filas_leidas += 1
            valores = list(fila)

            if indice_fila >= 2:
                valor_columna_d = valores[3] if len(valores) >= 4 else None
                if debe_eliminar(valor_columna_d):
                    continue

            writer.writerow(["" if valor is None else valor for valor in valores])
            filas_escritas += 1

    workbook.close()
    return filas_leidas, filas_escritas


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Filtra ARTS_ICEPAR.xlsx eliminando filas cuya columna D sea 0 o '-' y exporta a CSV."
    )
    parser.add_argument(
        "entrada",
        nargs="?",
        default="ARTS_ICEPAR.xlsx",
        help="Ruta del archivo Excel de entrada.",
    )
    parser.add_argument(
        "salida",
        nargs="?",
        default="ARTS_ICEPAR.csv",
        help="Ruta del archivo CSV de salida.",
    )
    parser.add_argument(
        "--hoja",
        dest="hoja",
        default=None,
        help="Nombre de la hoja a procesar. Si no se indica, se usa la hoja activa.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    archivo_entrada = Path(args.entrada).resolve()
    archivo_salida = Path(args.salida).resolve()

    if not archivo_entrada.exists():
        raise FileNotFoundError(f"No se encontró el archivo de entrada: {archivo_entrada}")

    filas_leidas, filas_escritas = procesar_excel_a_csv(
        archivo_entrada=archivo_entrada,
        archivo_salida=archivo_salida,
        hoja=args.hoja,
    )

    print(f"Archivo generado: {archivo_salida}")
    print(f"Filas leídas: {filas_leidas}")
    print(f"Filas escritas: {filas_escritas}")
    print(f"Filas eliminadas: {filas_leidas - filas_escritas}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())